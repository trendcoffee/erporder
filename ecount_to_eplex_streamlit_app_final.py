
import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# 하드코딩된 매핑
elevenst_map = {
    '1883시럽모음:1883바닐라': '3651744373_1',
    '1883시럽모음:1883카라멜': '3651744373_2',
    '1883시럽모음:1883로스티드헤이즐넛': '3651744373_3',
    '1883시럽모음:1883그린민트': '3651744373_4',
    '1883시럽모음:1883블루큐라소': '3651744373_5',
    '1883시럽모음:1883레몬': '3651744373_6',
    '1883시럽모음:1883모히또': '3651744373_7',
    '1883시럽모음:1883초콜릿': '3651744373_8',
    '1883시럽모음:1883아이스티피치': '3651744373_9',
    '1883시럽모음:1883스트로베리': '3651744373_10',
    '1883시럽모음:1883오렌지': '3651744373_11',
    '1883시럽모음:1883키위': '3651744373_12',
    '1883시럽모음:1883자몽': '3651744373_13',
    '1883시럽모음:1883핑크자몽': '3651744373_14',
    '1883시럽모음:1883패션프룻': '3651744373_15',
    '1883시럽모음:1883망고': '3651744373_16',
    '1883시럽모음:1883라임': '3651744373_17',
    '1883시럽모음:1883로즈': '3651744373_18',
    '1883시럽모음:1883애플': '3651744373_19',
    '1883시럽모음:1883바나나': '3651744373_20',
    '1883시럽모음:1883블루베리': '3651744373_21',
    '1883시럽모음:1883체리': '3651744373_22',
    '1883시럽모음:1883케인슈가': '3651744373_23',
    '1883시럽모음:1883피치': '3651744373_24',
    '1883시럽모음:1883차이티': '3651744373_25',
    '1883시럽모음:1883솔티드카라멜': '3651744373_26',
    '1883시럽모음:1883시나몬': '3651744373_27',
    '1883시럽모음:1883라벤더': '3651744373_28',
    '1883시럽모음:1883화이트초코': '3651744373_29',
    '1883시럽모음:1883석류': '3651744373_30',
    '1883시럽모음:1883라즈베리': '3651744373_31',
    '1883시럽모음:1883파인애플': '3651744373_32',
    '1883시럽모음:1883아이리쉬크림': '3651744373_33',
    '1883시럽모음:1883그린애플': '3651744373_34',
    '1883시럽모음:1883돌체드레체': '3651744373_35',
    '1883시럽모음:1883엘더플라워': '3651744373_36',
    '1883시럽모음:1883마카다미아넛': '3651744373_37',
    '1883시럽모음:1883아몬드': '3651744373_38',
    '1883시럽모음:1883리치': '3651744373_39',
    '1883시럽모음:1883화이트피치': '3651744373_40',
    '1883시럽모음:1883시럽펌프2개': '3651744373_41'
}

lotteon_map = {
    '바닐라시럽1000ml': 'LO1506416845_1',
    '카라멜시럽1000ml': 'LO1506416845_2',
    '헤이즐넛시럽1000ml': 'LO1506416845_3',
    '그린민트시럽1000ml': 'LO1506416845_4',
    '블루큐라소시럽1000ml': 'LO1506416845_5',
    '레몬시럽1000ml': 'LO1506416845_6',
    '모히또시럽1000ml': 'LO1506416845_7',
    '초콜릿시럽1000ml': 'LO1506416845_8',
    '아이스티피치시럽1000ml': 'LO1506416845_9',
    '스트로베리시럽1000ml': 'LO1506416845_10',
    '오렌지시럽1000ml': 'LO1506416845_11',
    '키위시럽1000ml': 'LO1506416845_12',
    '자몽시럽1000ml': 'LO1506416845_13',
    '핑크자몽시럽1000ml': 'LO1506416845_14',
    '패션프릇시럽1000ml': 'LO1506416845_15',
    '망고시럽1000ml': 'LO1506416845_16',
    '라임시럽1000ml': 'LO1506416845_17',
    '로즈시럽1000ml': 'LO1506416845_18',
    '애플시럽1000ml': 'LO1506416845_19',
    '바나나시럽1000ml': 'LO1506416845_20',
    '블루베리시럽1000ml': 'LO1506416845_21',
    '체리시럽1000ml': 'LO1506416845_22',
    '케인슈가시럽1000ml': 'LO1506416845_23',
    '피치시럽1000ml': 'LO1506416845_24',
    '차이티시럽1000ml': 'LO1506416845_25',
    '솔티드카라멜시럽1000ml': 'LO1506416845_26',
    '시나몬시럽1000ml': 'LO1506416845_27',
    '라벤더시럽1000ml': 'LO1506416845_28',
    '화이트초코시럽1000ml': 'LO1506416845_29',
    '석류시럽1000ml': 'LO1506416845_30',
    '라즈베리시럽1000ml': 'LO1506416845_31',
    '파인애플시럽1000ml': 'LO1506416845_32',
    '아이리쉬크림시럽1000ml': 'LO1506416845_33',
    '그린애플시럽1000ml': 'LO1506416845_34',
    '돌체드레체시럽1000ml': 'LO1506416845_35',
    '엘더플라워시럽1000ml': 'LO1506416845_36',
    '1883시럽펌프': 'LO1506416845_37',
    '리치시럽1000ml': 'LO1506416845_38',
    '화이트피치시럽1000ml': 'LO1506416845_39',
    '아몬드시럽1000ml': 'LO1506416845_40',
    '마카다미아넛시럽1000ml': 'LO1506416845_41',
    '': 'LO1506416845_42'
}

# 여기에 변환 함수 등 이어서 붙일 예정


def convert_excel(order_df, bom_df):
    today = datetime.today().strftime('%Y-%m-%d')
    order_df = order_df.iloc[:-1]
    elevenst_clean = {k.replace(' ', ''): v for k, v in elevenst_map.items()}
    lotteon_clean = {k.replace(' ', ''): v for k, v in lotteon_map.items()}
    rows = []

    for _, row in order_df.iterrows():
        try:
            주문번호 = str(int(float(row.get('주문번호', ''))))
            묶음주문번호 = str(int(float(row.get('묶음주문번호', ''))))
        except:
            주문번호, 묶음주문번호 = '', ''

        우편번호 = str(row.get('우편번호', '')).split('.')[0].zfill(5)
        수집처 = str(row['수집처']).strip()
        옵션 = str(row.get('주문옵션', '')).strip().replace(' ', '')
        key = str(row.get('쇼핑몰품목key', '')).strip()
        erp = str(row.get('품목코드(ERP)', '')).strip()

        if 수집처 == '쿠팡':
            code = 묶음주문번호
        elif 수집처 == '11번가':
            if 옵션 and 옵션 in elevenst_clean:
                code = elevenst_clean[옵션]
            else:
                code = key.split('/')[0]
        elif 수집처 == '롯데ON':
            if 옵션 and 옵션 in lotteon_clean:
                code = lotteon_clean[옵션]
            elif key.startswith('LO'):
                code = key.split('/')[0]
            else:
                code = erp
        else:
            code = erp

        is_missing = code not in bom_df['판매품목코드'].astype(str).values
        if 수집처 == '롯데ON' and code != 'UNKNOWN':
            is_missing = False

        rows.append({
            '* F/C': 'NS001',
            '* 주문유형': '7',
            '* 배송처': '17',
            '* 고객ID': '90015746',
            '판매채널': 수집처,
            '* 묶음배송번호': 주문번호,
            '* 품목코드': code,
            '품목명': row['품목명(ERP)'],
            '옵션': '',
            '가격': row['주문금액'],
            '* 품목수량': row['수량'],
            '주문자': row['주문자'],
            '* 받는사람명': row['수취인'],
            '주문자 전화번호': row['주문자연락처'],
            '* 받는사람 전화번호': row['수취인연락처1'],
            '* 받는사람 우편번호': 우편번호,
            '* 받는사람 주소': row['주소'],
            '배송메세지': row['배송요청사항'],
            '* 주문일자': today,
            '상품주문번호': '',
            '주문번호(참조)': 묶음주문번호,
            '박스구분': '',
            '상세배송유형': '',
            '새벽배송 SMS 전송': '',
            '새벽배송 현관비밀번호': '',
            '위험물 구분': '',
            '* 주문중개채널': 'SELF',
            'API 연동용 판매자ID': '',
            '* 주문시간': '09:00:00',
            '받는사람 핸드폰': '',
            '_ERROR': is_missing
        })

    return pd.DataFrame(rows)

st.title("이카운트 → 이플렉스 양식 변환기 (Styled Streamlit)")

ecount_file = st.file_uploader("① 이카운트 주문양식 업로드", type=["xlsx"])
bom_file = st.file_uploader("② BOM 등록 리스트 업로드", type=["csv"])

if ecount_file and bom_file:
    order_df = pd.read_excel(ecount_file, skiprows=1)
    bom_df = pd.read_csv(bom_file)

    if st.button("변환 실행"):
        df = convert_excel(order_df, bom_df)
        df_show = df.drop(columns=['_ERROR'])
        st.success("변환 완료! 결과를 아래에서 확인하거나 다운로드하세요.")
        st.dataframe(df_show)

        # 파일 저장 + 스타일 적용
        temp_file = "temp_styled_output.xlsx"
        df_show.to_excel(temp_file, index=False)
        wb = load_workbook(temp_file)
        ws = wb.active
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        for i, err in enumerate(df['_ERROR'], start=2):
            if err:
                for j in range(1, ws.max_column + 1):
                    ws.cell(row=i, column=j).fill = red_fill

        styled_file = "styled_output.xlsx"
        wb.save(styled_file)

        with open(styled_file, "rb") as f:
            st.download_button("📥 변환 결과 다운로드", f, file_name="이플렉스_변환결과.xlsx")
